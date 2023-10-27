from openpyxl import *
from openpyxl.utils import *
from functions import *
from datetime import datetime
import builtins

def addRows():
    with builtins.open("input.txt", 'r') as file:
        lines = file.readlines()

    filename = lines[0].strip()
    endOfLastMonth = lines[1].strip()
    endOfThisMonth = lines[2].strip()
    dateToProcess = lines[3].strip()

    try:
        dateToProcess = datetime.strptime(dateToProcess, '%Y-%m-%d')
    except ValueError:
        print("Invalid date format. Please enter the date in yyyy-mm-dd format.")
        exit()

    wb = load_workbook(filename)
    ws = wb.active

    rowsToInsert = []
    clientRows = []

    # Find location of rows where new row will go after
    for row in ws.iter_rows(min_col=2, max_col=2):
        for cell in row:
            prevRow = cell.row - 1
            nextRow = cell.row + 1

            if (cell.value == endOfLastMonth) and (ws.cell(row=nextRow, column=2).value == "Annual Bonus 4.5%"):
                rowsToInsert.append(cell.row + 1)
                continue
            
            if cell.value == endOfLastMonth:
                rowsToInsert.append(cell.row)
            
    # Find location of client rows
    # Client rows are used to determine IR/BR, as well as to extract client information later on
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "":
                continue

            if cell.value is not None and cell.value != "Subitems":
                clientRow = cell.row
                clientRows.append(clientRow)

    # Initialize indexing for clientRows
    # The value to subtract by is affected by the number of new clients who do not have subitems
    cr = len(clientRows) - 2

    # Loop through rowsToInsert and insert a new row after the current row
    # Iteratively update clientRows as the spreadsheet is updated with new rows
    for row in reversed(rowsToInsert):
        ws.insert_rows(row + 1)
        newRow = row + 1

        clientRows = [r + 1 if r >= row + 1 else r for r in clientRows] # Update clientRows as newRows are added

        copyFormat(ws, newRow - 1, 1, newRow, 1)

        # Copy over Name, Status, Beginning Balance, Interest Earned
        # TODO: Add feature to add a "Reinvestment" or "Interest Payout" row after "Annual Bonus",
        # depending on IR/BR status
        ws.cell(row=newRow, column=2, value=endOfThisMonth)
        ws.cell(row=newRow, column=3, value=ws.cell(row=newRow - 1, column=3).value) # Reinvestment/Interest Payout/Annual Bonus
        ws.cell(row=newRow, column=4, value=ws.cell(row=newRow - 1, column=6).value.replace("$", "").replace(",", ""))
        ws.cell(row=newRow, column=5, value=ws.cell(row=newRow - 1, column=7).value)

        copyFormat(ws, newRow - 1, 2, newRow, 2)
        copyFormat(ws, newRow - 1, 3, newRow, 3)
        copyFormat(ws, newRow - 1, 4, newRow, 4)
        copyFormat(ws, newRow - 1, 5, newRow, 5)

        # Calculate and copy over Ending Balance, Following Month Accrual, Date to Process
        beginningBal = float(ws.cell(row=newRow, column=4).value)
        interestEarned = float(ws.cell(row=newRow, column=5).value)

        # If IR is checked, add Beginning Bal + Interest Earned
        # If IR is NOT checked, Ending Bal = Beginning Bal
        if cr <= len(clientRows):
            if (ws.cell(row=clientRows[cr], column=15).value is not None and ws.cell(row=clientRows[cr], column=15).value != " ") and ws.cell(row=clientRows[cr], column=15).value == "v":
                endingBal = beginningBal + interestEarned
                endingBal = "${:,.2f}".format(endingBal)
                endingBalCleaned = endingBal.replace("$", "").replace(",", "")

                followingMonthAccrual = nextMonthInterest(float(endingBalCleaned))

                ws.cell(row=newRow, column=6, value=endingBal)
                ws.cell(row=newRow, column=7, value=followingMonthAccrual)  
                ws.cell(row=newRow, column=8, value=dateToProcess)

                copyFormat(ws, newRow - 1, 6, newRow, 6)
                copyFormat(ws, newRow - 1, 7, newRow, 7)
                copyFormat(ws, newRow - 1, 8, newRow, 8)

                # If IR, copy interestEarned to Interest Reinvested column
                # and move previous Interest Paid Out value down
                ws.cell(row=newRow, column=9, value=interestEarned)
                ws.cell(row=newRow, column=10, value=ws.cell(row=newRow - 1, column=10).value)

                copyFormat(ws, newRow - 1, 9, newRow, 9)
                copyFormat(ws, newRow - 1, 10, newRow, 10)

                cr -= 1
            else:
                endingBal = beginningBal
                endingBal = "${:,.2f}".format(endingBal)
                endingBalCleaned = endingBal.replace("$", "").replace(",", "")

                followingMonthAccrual = nextMonthInterest(float(endingBalCleaned))

                ws.cell(row=newRow, column=6, value=endingBal)
                ws.cell(row=newRow, column=7, value=followingMonthAccrual)   
                ws.cell(row=newRow, column=8, value=dateToProcess)
    
                copyFormat(ws, newRow - 1, 6, newRow, 6)
                copyFormat(ws, newRow - 1, 7, newRow, 7)
                copyFormat(ws, newRow - 1, 8, newRow, 8)

                # If NOT IR, copy interestEarned to Interest Paid Out column
                # and move previous Interest Reinvested value down
                ws.cell(row=newRow, column=10, value=interestEarned)
                ws.cell(row=newRow, column=9, value=ws.cell(row=newRow - 1, column=9).value)

                copyFormat(ws, newRow - 1, 9, newRow, 9)
                copyFormat(ws, newRow - 1, 10, newRow, 10)

                cr -= 1
        
    wb.save(lines[4].strip())

addRows()



