from openpyxl import *
from openpyxl.utils import *
from functions import *
import builtins

def calcTotalInterest():
    with builtins.open("input.txt", 'r') as file:
        lines = file.readlines()

    filename = lines[4].strip()

    wb = load_workbook(filename)
    ws = wb.active

    # Adds up Interest Reinvested column for each client and assigns 
    # result to Total Interest Reinvested column in Client Row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "":
                continue

            if cell.value is not None and cell.value != "Subitems":
                clientRow = cell.row
                lookupRow = cell.row + 2
                interestReinvestedSum = 0

                while ws.cell(row=lookupRow, column=1).value == "" or ws.cell(row=lookupRow, column=1).value is None:
                    if ws.cell(row=lookupRow, column=9).value is not None:
                        interestReinvestedSum += ws.cell(row=lookupRow, column=9).value

                    if ws.cell(row=lookupRow + 1, column=1).value is not None:
                        break

                    if lookupRow > 10000:
                        break
                    
                    if lookupRow <= 10000:
                        lookupRow += 1

                ws.cell(row=clientRow, column=19, value=interestReinvestedSum)

    # Adds up Interest Paid Out column for each client and assigns 
    # result to Total Interest Paid Out column in Client Row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "":
                continue

            if cell.value is not None and cell.value != "Subitems":
                clientRow = cell.row
                lookupRow = cell.row + 2
                interestPaidOutSum = 0

                while ws.cell(row=lookupRow, column=1).value == "" or ws.cell(row=lookupRow, column=1).value is None:
                    if ws.cell(row=lookupRow, column=10).value is not None:
                        interestPaidOutSum += ws.cell(row=lookupRow, column=10).value

                    if ws.cell(row=lookupRow + 1, column=1).value is not None:
                        break

                    if lookupRow > 10000:
                        break
                    
                    if lookupRow <= 10000:
                        lookupRow += 1

                ws.cell(row=clientRow, column=20, value=interestPaidOutSum)

    # Adds up Total Int. Reinvested + Total Int. Paid Out and assigns the
    # result to Total Interest Earned column in Client Row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "":
                continue

            if cell.value is not None and cell.value != "Subitems":
                clientRow = cell.row
                interestEarnedTotal = ws.cell(row=clientRow, column=19).value + ws.cell(row=clientRow, column=20).value
                ws.cell(row=clientRow, column=18, value=interestEarnedTotal)

    wb.save(lines[5].strip())

calcTotalInterest()