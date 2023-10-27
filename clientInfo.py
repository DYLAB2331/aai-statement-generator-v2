from openpyxl import *
from openpyxl.utils import *
from functions import *
from pdfGen import generatePDF, drawFooter, drawHeader
import builtins

def getClientInfo():
    with builtins.open("input.txt", 'r') as file:
        lines = file.readlines()

        fileName = lines[5].strip()
        endOfThisMonth = lines[2].strip()

        wb = load_workbook(fileName)
        ws = wb.active

        clientRows = []
        clientInfo = {}
        clientFreq = {}
        currentCount = {}
        endingBalRows = []
        endingBalIndex = 0
        clientTransactionHistory  = [[]]
        clientIndex = 0

        # Find rows containing client information
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == "":
                    continue

                if cell.value is not None and cell.value != "Subitems":
                    clientRow = cell.row
                    clientRows.append(clientRow)

        # Append client transaction history to client position in array
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == "Subitems":
                    continue

                if cell.value is not None and cell.value != "" and cell.value != "Subitems":
                    clientRow = cell.row
                    lookupRow = cell.row + 2

                    while ws.cell(row=lookupRow, column=1).value == "" or ws.cell(row=lookupRow, column=1).value is None:                            
                        transactionRow = []
                        transactionRow.extend([
                            ws.cell(row=lookupRow, column=2).value,
                            ws.cell(row=lookupRow, column=3).value,
                            ws.cell(row=lookupRow, column=4).value,
                            ws.cell(row=lookupRow, column=5).value,
                            ws.cell(row=lookupRow, column=6).value,
                        ])

                        clientTransactionHistory[clientIndex].append(transactionRow)

                        if ws.cell(row=lookupRow + 1, column=1).value is not None:
                            break

                        if lookupRow > 10000:
                            break
                        
                        if lookupRow <= 10000:
                            lookupRow += 1

                    clientIndex += 1
                    clientTransactionHistory.append([])

        # Find rows containing endOfThisMonth
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                prevRow = cell.row - 1
                nextRow = cell.row + 1

                if (cell.value == endOfThisMonth) and (ws.cell(row=nextRow, column=2).value == "Annual Bonus 4.5%"):
                    endingBalRows.append(cell.row + 1)
                    continue
                
                if cell.value == endOfThisMonth:
                    endingBalRows.append(cell.row)

        # Loop through each clientRow and get the info of each client
        # Append info to infoList and store list as a value with account number acting as key
        for clientRow in clientRows:
            infoList = []

            infoList.append(ws.cell(row=clientRow, column=1).value) # Name
            infoList.append(ws.cell(row=clientRow, column=7).value) # Account Type TODO: Convert Cash = Non-Retirement
            infoList.append(ws.cell(row=clientRow, column=8).value) # Funded Date 
            infoList.append(ws.cell(row=clientRow, column=5).value) # Principal
            infoList.append(ws.cell(row=clientRow, column=11).value) # Monthly Interest
            infoList.append(ws.cell(row=clientRow, column=14).value) # Annual Bonus
            infoList.append(ws.cell(row=clientRow, column=15).value) # IR
            infoList.append(ws.cell(row=clientRow, column=16).value) # BR
            infoList.append(ws.cell(row=clientRow, column=18).value) # Total Interest Earned

            endingBalEOM = ws.cell(row=endingBalRows[endingBalIndex], column=6).value
            endingBalIndex += 1
            infoList.append(endingBalEOM) # Index 9 in infoList

            infoList.append(ws.cell(row=clientRow, column=6).value) # Duration, index 10

            acctNo = ws.cell(row=clientRow, column=4).value

            clientInfo[acctNo] = infoList

        # Print dict for testing purposes
        for key, value in clientInfo.items():
            print(f"{key}: ", end='')
            if isinstance(value, list):
                for item in value:
                    print(item, end=', ')
                print()  
            else:
                print(value)

        # Iterate through clientInfo dictionary and count the number of occurences of a client's name
        # Store the frequency of a client name in clientFreq with their name as the key
        for valueList in clientInfo.values():
            if valueList[0] in clientFreq:
                clientFreq[valueList[0]] += 1
            else:
                clientFreq[valueList[0]] = 1

        # Print dict for testing purposes
        for key, value in clientFreq.items():
            print(f"{key}: ", end='')
            if isinstance(value, list):
                for item in value:
                    print(item, end=', ')
                print()  
            else:
                print(value)

        # Reset clientIndex
        clientIndex = 0

        # Iterate through clientInfo dictionary and generate PDF for specific client
        for key, valueList in clientInfo.items():
            clientName = valueList[0]

            if clientName not in currentCount:
                currentCount[clientName] = 0
            print(clientIndex)
            currentCount[clientName] += 1
            generatePDF(key, valueList, clientFreq, currentCount[clientName], clientTransactionHistory[clientIndex])
            clientIndex += 1

getClientInfo()





